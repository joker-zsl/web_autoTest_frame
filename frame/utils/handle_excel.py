# -*- coding:utf-8 -*-
# @Time : 2021/5/11
# @Author : 张顺利
# @Group : 云业务测试部
# @Email : zhangshunl@yuanian.com

import openpyxl


class ReadExcel:
    """读取excel文件"""
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname
        self.workbook = None
        self.sheet = None
        self.value = None

    def open(self):
        self.workbook = openpyxl.load_workbook(self.filename)
        self.sheet = self.workbook[self.sheetname]

    def close(self):
        self.workbook.close()

    def data(self):
        self.open()
        max_row = self.sheet.max_row
        max_column = self.sheet.max_column
        data_list = []
        for row in range(1, max_row + 1):
            row_data = []
            for column in range(1, max_column + 1):
                value = self.sheet.cell(row=row, column=column).value
                row_data.append(value)
            data_list.append(row_data)

        dict_data_list = []
        title = data_list[0]
        for data in data_list[1:]:
            dict_data = dict(zip(title, data))
            dict_data_list.append(dict_data)
        self.close()
        self.value = dict_data_list
        return self

    def filter_by(self, **kwargs):
        """过滤出分别满足条件的数据"""
        datas = self.value
        filter_data = []
        for key, value in kwargs.items():
            for data in datas:
                if data[key] == value:
                    datas.remove(data)
                    filter_data.append(data)
        return filter_data
